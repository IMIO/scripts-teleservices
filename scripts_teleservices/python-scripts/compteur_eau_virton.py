import csv
import openpyxl
from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
from odf.text import P
from datetime import datetime

# --- Fonction ajoutée pour corriger le problème des zéros ---
def normalize_id(value):
    """
    Convertit la valeur en chaîne, retire les espaces et les zéros au début.
    Exemple: "00123" devient "123", "123" reste "123".
    """
    if value is None:
        return ""
    # On convertit en string et on nettoie les espaces
    s = str(value).strip()
    # On retire les '0' à gauche. 
    # Attention: si la valeur est juste "0", cela renvoie une chaîne vide, 
    # ce qui est acceptable tant que c'est cohérent pour les deux fichiers.
    return s.lstrip("0")

# Fonction pour lire les données de la feuille XLSX
def read_xlsx(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

# Fonction pour lire les données de la feuille ODS
def read_ods(sheet):
    data = []
    for row in sheet.getElementsByType(TableRow):
        row_data = []
        for cell in row.getElementsByType(TableCell):
            cell_text = "".join(str(text) for text in cell.getElementsByType(P))
            row_data.append(cell_text)
        data.append(row_data)
    return data

# Fonction pour formater les dates
def format_date_to_clean(date_str):
    if isinstance(date_str, datetime):
        return date_str.strftime("%d/%m/%Y")
    try:
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
    except ValueError:
        return date_str  # Retourner la date originale si le format ne correspond pas

def format_date_from_ods(date_str):
    try:
        # Essayer de parser si c'est une chaine simple
        return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        # Parfois ODS renvoie déjà un format iso ou autre, on renvoie tel quel
        return date_str

# --- Chargement des fichiers ---
# Assurez-vous que les noms de fichiers correspondent à vos fichiers réels
wb1 = openpyxl.load_workbook("Extraction relevés 2025(1).xlsx") # Modifiez si nécessaire
sheet1 = wb1.active

ods = load("releve-de-compteurs.ods") # Modifiez si nécessaire
sheet2 = ods.spreadsheet.getElementsByType(Table)[0]

data1 = read_xlsx(sheet1)
data2 = read_ods(sheet2)

# Afficher les en-têtes pour débogage
print("Headers from fichier1 (XLSX):", data1[0])
print("Headers from fichier2 (ODS):", data2[0])

# Extraire les en-têtes
headers1 = data1[0]
headers2 = data2[0]

# Vérifier la présence de "MATRICULE" dans les en-têtes et corriger si nécessaire
matricule_header_ods = "N° de matricule" 

if "MATRICULE" not in headers1:
    raise ValueError("'MATRICULE' is not in the headers of fichier1.xlsx")
if matricule_header_ods not in headers2:
    raise ValueError(f"'{matricule_header_ods}' is not in the headers of fichier2.ods")

# Renommer "N° de matricule" en "MATRICULE" dans les données ODS pour homogénéiser
headers2 = ["MATRICULE" if h == matricule_header_ods else h for h in headers2]
data2[0] = headers2

# --- Modification ici : Normalisation des clés pour le dictionnaire ---
index1_matricule = headers1.index("MATRICULE")
index1_compteur = headers1.index("NO_COMPTEUR")

# On utilise normalize_id sur la clé (matricule, compteur)
data1_dict = {
    (normalize_id(row[index1_matricule]), normalize_id(row[index1_compteur])): row 
    for row in data1[1:]
}

# Définir la correspondance des colonnes
correspondance_colonnes = {
    "NOUVEL_INDEX": "Index du compteur",
    "CONSOMMATION": "Consommation",
    "DATE_RELEVE": "Date",
    "TYPE_CONSOMMATION": "Consommation non-ordinnaire",
    "NOM": "Nom",
    "PRENOM": "Prénom",
    "ANCIEN_INDEX": "Index précédent",
    "DATE_ANCIEN_RELEVE": "Date du précédent relevé",
}

# Créer une liste pour stocker les données mises à jour
updated_data1 = [headers1]
clean_data = [headers1]

index2_matricule = headers2.index("MATRICULE")
index2_compteur = headers2.index("N° de compteur")

# Mettre à jour les valeurs dans data1 avec les valeurs correspondantes de data2
for row2 in data2[1:]:
    # --- Modification ici : Normalisation des clés pour la recherche ---
    matricule2_norm = normalize_id(row2[index2_matricule])
    no_compteur2_norm = normalize_id(row2[index2_compteur])
    
    key_lookup = (matricule2_norm, no_compteur2_norm)

    if key_lookup in data1_dict:
        row1 = data1_dict[key_lookup]
        row_modified = False
        for col_name1, col_name2 in correspondance_colonnes.items():
            if col_name2 in headers2:
                idx1 = headers1.index(col_name1)
                idx2 = headers2.index(col_name2)
                
                # Vérifier qu'on ne sort pas des limites de la ligne ODS
                if idx2 < len(row2) and row2[idx2]:
                    val_ods = row2[idx2]
                    if col_name1 in ["DATE_RELEVE", "DATE_ANCIEN_RELEVE"]:
                        formatted_date = format_date_from_ods(val_ods)
                        row1[idx1] = formatted_date
                    else:
                        row1[idx1] = val_ods
                    row_modified = True
        
        if row_modified:
            clean_row = [
                format_date_to_clean(item) if headers1[i] in ["DATE_RELEVE", "DATE_ANCIEN_RELEVE"] else item
                for i, item in enumerate(row1)
            ]
            clean_data.append(clean_row)
        updated_data1.append(row1)
    else:
        # Cas où la ligne ODS n'existe pas dans le XLSX
        new_row = [""] * len(headers1)
        # On remplit avec les infos qu'on a
        new_row[index1_matricule] = row2[index2_matricule] # On garde la valeur originale non tronquée si possible, ou celle de l'ODS
        new_row[index1_compteur] = row2[index2_compteur]
        
        for col_name1, col_name2 in correspondance_colonnes.items():
            if col_name2 in headers2:
                idx1 = headers1.index(col_name1)
                idx2 = headers2.index(col_name2)
                if idx2 < len(row2) and row2[idx2]:
                    if col_name1 in ["DATE_RELEVE", "DATE_ANCIEN_RELEVE"]:
                        formatted_date = format_date_from_ods(row2[idx2])
                        new_row[idx1] = formatted_date
                    else:
                        new_row[idx1] = row2[idx2]
        
        clean_row = [
            format_date_to_clean(item) if headers1[i] in ["DATE_RELEVE", "DATE_ANCIEN_RELEVE"] else item
            for i, item in enumerate(new_row)
        ]
        clean_data.append(clean_row)
        updated_data1.append(new_row)

# Ajouter les lignes de data1 qui ne sont pas dans data2
# Pour optimiser, on crée d'abord un ensemble (set) des clés présentes dans le fichier ODS
keys_in_ods = set()
for row in data2[1:]:
    m = normalize_id(row[index2_matricule])
    c = normalize_id(row[index2_compteur])
    keys_in_ods.add((m, c))

for key, row in data1_dict.items():
    # 'key' est déjà normalisée ici car data1_dict a été construit avec normalize_id
    if key not in keys_in_ods:
        updated_data1.append(row)

# Écrire les données mises à jour dans un nouveau fichier XLSX
wb_out = openpyxl.Workbook()
sheet_out = wb_out.active

for row in updated_data1:
    sheet_out.append(row)

wb_out.save("fichier_merge.xlsx")
print("Fichier fusionné enregistré sous 'fichier_merge.xlsx'")

# Écrire les données modifiées dans un nouveau fichier "clean"
if len(clean_data) > 1: # Plus que juste les en-têtes
    wb_clean = openpyxl.Workbook()
    sheet_clean = wb_clean.active

    for row in clean_data:
        sheet_clean.append(row)

    wb_clean.save("fichier_clean.xlsx")
    print("Fichier nettoyé enregistré sous 'fichier_clean.xlsx'")
else:
    print("Aucune donnée modifiée trouvée.")